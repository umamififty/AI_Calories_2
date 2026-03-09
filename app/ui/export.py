"""
Standalone Tkinter export window — Claude Dark theme.
Launched as subprocess by menubar_app.py.
Exports consumption_log to CSV or Excel on the Desktop.
"""
import tkinter as tk
from tkinter import ttk, messagebox
import csv
import os
import sqlite3
from datetime import date, timedelta, datetime

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "data", "nutrition.db")
DESKTOP = os.path.expanduser("~/Desktop")

COLUMNS = ["date", "food_name", "calories", "protein", "fat", "carbs", "sugar", "fiber", "sodium", "logged_at"]
HEADERS = ["Date", "Food", "Calories (kcal)", "Protein (g)", "Fat (g)", "Carbs (g)", "Sugar (g)", "Fiber (g)", "Sodium (mg)", "Logged At"]

# ── Claude Dark Palette ────────────────────────────────────────
BG          = "#1E1B18"
SURFACE     = "#272320"
INPUT_BG    = "#201D1A"
BORDER      = "#3D3830"
ACCENT      = "#C96B3B"
ACCENT_HOV  = "#A8552D"
TEXT        = "#F0EDE8"
SUBTEXT     = "#9E9A94"
UNIT_TEXT   = "#6B6560"
SECT_TEXT   = "#C96B3B"
BTN_BG      = "#2C2825"
BTN_HOV     = "#38332E"

FONT_HEAD   = ("Helvetica Neue", 17, "bold")
FONT_LABEL  = ("Helvetica Neue", 13)
FONT_ENTRY  = ("Helvetica Neue", 13)
FONT_UNIT   = ("Helvetica Neue", 11)
FONT_BTN    = ("Helvetica Neue", 12, "bold")
FONT_BTN_SM = ("Helvetica Neue", 11)
FONT_SECT   = ("Helvetica Neue", 9, "bold")


def make_btn(parent, text, command, primary=False) -> tk.Button:
    bg  = ACCENT     if primary else BTN_BG
    hov = ACCENT_HOV if primary else BTN_HOV
    fg  = "#FFFFFF"  if primary else TEXT
    font = FONT_BTN  if primary else FONT_BTN_SM

    b = tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=fg, activebackground=hov, activeforeground=fg,
        relief="flat", borderwidth=0, padx=16, pady=7,
        font=font, cursor="hand2",
    )
    b.bind("<Enter>", lambda _: b.config(bg=hov))
    b.bind("<Leave>", lambda _: b.config(bg=bg))
    return b


def fetch_rows(start_date: str, end_date: str) -> list:
    if not os.path.exists(DB_PATH):
        return []
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM consumption_log WHERE date >= ? AND date <= ? ORDER BY date, logged_at",
        (start_date, end_date),
    )
    rows = cursor.fetchall()
    conn.close()
    result = []
    for row in rows:
        d = dict(row)
        result.append([d.get(col, "") for col in COLUMNS])
    return result


def export_csv(path: str, rows: list):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(rows)


def export_excel(path: str, rows: list):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nutrition Log"

    header_fill = PatternFill("solid", fgColor="C96B3B")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)


def build_window():
    root = tk.Tk()
    root.title("Export Nutrition Data")
    root.resizable(False, False)
    root.configure(bg=BG)

    # ── Header ────────────────────────────────────────────────
    header = tk.Frame(root, bg=BG)
    header.pack(fill="x", padx=0, pady=0)

    tk.Frame(header, bg=ACCENT, width=4).pack(side="left", fill="y")

    header_inner = tk.Frame(header, bg=BG)
    header_inner.pack(side="left", fill="x", expand=True, padx=(16, 20), pady=(20, 16))

    tk.Label(header_inner, text="Export Data",
             bg=BG, fg=TEXT, font=FONT_HEAD).pack(anchor="w")
    tk.Label(header_inner, text="Save your nutrition log to CSV or Excel",
             bg=BG, fg=SUBTEXT, font=FONT_UNIT).pack(anchor="w", pady=(2, 0))

    # ── Divider ───────────────────────────────────────────────
    tk.Frame(root, bg=BORDER, height=1).pack(fill="x")

    content = tk.Frame(root, bg=BG)
    content.pack(fill="both", expand=True, padx=24, pady=(16, 0))

    # ── Date Range ────────────────────────────────────────────
    sect_frame = tk.Frame(content, bg=BG)
    sect_frame.pack(fill="x", pady=(8, 6))
    tk.Label(sect_frame, text="DATE RANGE", bg=BG, fg=SECT_TEXT, font=FONT_SECT).pack(side="left")
    tk.Frame(sect_frame, bg=BORDER, height=1).pack(
        side="left", fill="x", expand=True, padx=(10, 0), pady=(0, 2)
    )

    today = date.today()
    range_options = {
        "Today":        (today.isoformat(),                             today.isoformat()),
        "This Week":    ((today - timedelta(days=today.weekday())).isoformat(), today.isoformat()),
        "This Month":   (today.replace(day=1).isoformat(),              today.isoformat()),
        "Last 30 Days": ((today - timedelta(days=30)).isoformat(),      today.isoformat()),
        "All Time":     ("2000-01-01",                                  today.isoformat()),
    }

    range_var = tk.StringVar(value="This Week")

    range_row = tk.Frame(content, bg=SURFACE)
    range_row.pack(fill="x", pady=2)
    tk.Label(range_row, text="Date range", bg=SURFACE, fg=TEXT,
             font=FONT_LABEL, anchor="w", width=14).pack(side="left", padx=(14, 8), pady=9)

    # Style the combobox to fit the dark theme as closely as possible
    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("Dark.TCombobox",
        fieldbackground=INPUT_BG, background=INPUT_BG,
        foreground=TEXT, selectbackground=ACCENT, selectforeground=TEXT,
        bordercolor=BORDER, arrowcolor=TEXT, lightcolor=BORDER, darkcolor=BORDER,
    )
    style.map("Dark.TCombobox",
        fieldbackground=[("readonly", INPUT_BG)],
        foreground=[("readonly", TEXT)],
    )
    range_menu = ttk.Combobox(
        range_row, textvariable=range_var, values=list(range_options.keys()),
        state="readonly", width=14, font=FONT_ENTRY, style="Dark.TCombobox",
    )
    range_menu.pack(side="left", pady=9)

    # ── Format ────────────────────────────────────────────────
    sect_frame2 = tk.Frame(content, bg=BG)
    sect_frame2.pack(fill="x", pady=(14, 6))
    tk.Label(sect_frame2, text="FORMAT", bg=BG, fg=SECT_TEXT, font=FONT_SECT).pack(side="left")
    tk.Frame(sect_frame2, bg=BORDER, height=1).pack(
        side="left", fill="x", expand=True, padx=(10, 0), pady=(0, 2)
    )

    fmt_var = tk.StringVar(value="CSV")
    fmt_row = tk.Frame(content, bg=SURFACE)
    fmt_row.pack(fill="x", pady=2)

    for label, value in [("CSV", "CSV"), ("Excel (.xlsx)", "Excel")]:
        tk.Radiobutton(
            fmt_row, text=f"  {label}", variable=fmt_var, value=value,
            bg=SURFACE, fg=TEXT, selectcolor=BG,
            activebackground=SURFACE, activeforeground=TEXT,
            font=FONT_LABEL, borderwidth=0, relief="flat", cursor="hand2",
        ).pack(side="left", padx=(10, 16), pady=9)

    # ── Status label ──────────────────────────────────────────
    status_var = tk.StringVar(value="")
    tk.Label(content, textvariable=status_var, bg=BG, fg=SUBTEXT,
             font=FONT_UNIT).pack(anchor="w", pady=(10, 0))

    # ── Footer ────────────────────────────────────────────────
    tk.Frame(root, bg=BORDER, height=1).pack(fill="x", pady=(14, 0))

    btn_row = tk.Frame(root, bg=BG)
    btn_row.pack(fill="x", padx=24, pady=16)

    def on_export():
        selected = range_var.get()
        start, end = range_options[selected]
        rows = fetch_rows(start, end)

        if not rows:
            messagebox.showinfo(
                "No Data", f"No entries found for the selected range ({selected}).",
                parent=root,
            )
            return

        timestamp = datetime.now().strftime("%Y-%m-%d")
        fmt = fmt_var.get()

        if fmt == "CSV":
            path = os.path.join(DESKTOP, f"calories_export_{timestamp}.csv")
            export_csv(path, rows)
        else:
            path = os.path.join(DESKTOP, f"calories_export_{timestamp}.xlsx")
            try:
                export_excel(path, rows)
            except ImportError as e:
                messagebox.showerror("Missing Dependency", str(e), parent=root)
                return

        status_var.set(f"Exported {len(rows)} rows to {os.path.basename(path)}")
        messagebox.showinfo(
            "Export Complete", f"Saved to Desktop:\n{os.path.basename(path)}", parent=root
        )
        root.destroy()

    make_btn(btn_row, "Cancel", root.destroy).pack(side="left")
    make_btn(btn_row, "Export", on_export, primary=True).pack(side="right")

    # ── Center & bring to front ───────────────────────────────
    root.update_idletasks()
    w, h = root.winfo_reqwidth(), root.winfo_reqheight()
    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    root.lift()
    root.attributes("-topmost", True)
    root.after(500, lambda: root.attributes("-topmost", False))
    root.focus_force()

    root.mainloop()


if __name__ == "__main__":
    build_window()
